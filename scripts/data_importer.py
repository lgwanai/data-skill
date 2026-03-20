import argparse
import sqlite3
import pandas as pd
import os
import re
import warnings

# Suppress openpyxl warnings about data validation
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

def clean_column_names(columns):
    """Clean column names to be valid SQLite identifiers."""
    cleaned = []
    seen = set()
    for col in columns:
        if pd.isna(col):
            col = "Unnamed"
        
        # Convert to string and clean
        col_str = str(col).strip()
        # Replace non-alphanumeric with underscore
        col_str = re.sub(r'\W+', '_', col_str)
        # Remove leading/trailing underscores
        col_str = col_str.strip('_')
        
        if not col_str or col_str.isdigit():
            col_str = f"col_{col_str}"
            
        # Ensure unique
        final_col = col_str
        counter = 1
        while final_col.lower() in [s.lower() for s in seen]:
            final_col = f"{col_str}_{counter}"
            counter += 1
            
        seen.add(final_col)
        cleaned.append(final_col)
    return cleaned

def find_header_row(df, max_rows=10):
    """Find the most likely header row by looking for the row with the most non-null values."""
    best_row_idx = 0
    max_non_nulls = 0
    
    for i in range(min(max_rows, len(df))):
        non_null_count = df.iloc[i].count()
        if non_null_count > max_non_nulls:
            max_non_nulls = non_null_count
            best_row_idx = i
            
    return best_row_idx

def unmerge_and_fill_excel(file_path):
    """
    Reads an Excel file, unmerges any merged cells, and fills them with the top-left value.
    Returns a pandas DataFrame.
    """
    import openpyxl
    wb = openpyxl.load_workbook(file_path, data_only=False)
    sheet = wb.active
    
    # Extract merged cells ranges
    merged_ranges = list(sheet.merged_cells.ranges)
    
    # Iterate over merged cells and fill them with the top-left value
    for merged_cell in merged_ranges:
        min_col, min_row, max_col, max_row = merged_cell.bounds
        top_left_cell_value = sheet.cell(row=min_row, column=min_col).value
        
        # Unmerge the cells (optional, but good for clean structure)
        sheet.unmerge_cells(str(merged_cell))
        
        # Fill the unmerged cells with the top-left value
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                sheet.cell(row=row, column=col, value=top_left_cell_value)
                
    # Read into pandas
    data = sheet.values
    cols = next(data)
    df = pd.DataFrame(data, columns=cols)
    return df

def import_to_sqlite(file_path, db_path, table_name=None):
    """Import CSV/XLSX into SQLite, handling complex headers and merged cells."""
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"File not found: {file_path}")
        
    ext = os.path.splitext(file_path)[1].lower()
    base_name = os.path.splitext(os.path.basename(file_path))[0]
    
    if not table_name:
        # Clean base name for table name
        table_name = re.sub(r'\W+', '_', base_name).strip('_')
        
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    
    # Handle existing table by appending suffix
    original_table_name = table_name
    counter = 1
    while True:
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name=?", (table_name,))
        if not cursor.fetchone():
            break
        table_name = f"{original_table_name}_v{counter}"
        counter += 1
        
    print(f"Importing {file_path} to table '{table_name}' in {db_path}...")
    
    if ext == '.csv':
        # For CSV, use chunking for large files
        chunk_size = 50000
        first_chunk = True
        
        # Read a small sample to find the header
        sample_df = pd.read_csv(file_path, nrows=20, header=None)
        header_idx = find_header_row(sample_df)
        
        for chunk in pd.read_csv(file_path, skiprows=header_idx, chunksize=chunk_size):
            if first_chunk:
                chunk.columns = clean_column_names(chunk.columns)
                # Store columns for subsequent chunks
                final_cols = chunk.columns
                chunk.to_sql(table_name, conn, index=False, if_exists='replace')
                first_chunk = False
            else:
                chunk.columns = final_cols
                chunk.to_sql(table_name, conn, index=False, if_exists='append')
                
        print(f"CSV import completed successfully.")
        
    elif ext in ['.xlsx', '.xls']:
        try:
            # Try to handle merged cells first
            print("Processing Excel file (handling merged cells)...")
            df = unmerge_and_fill_excel(file_path)
            
            # Find real header
            header_idx = find_header_row(df)
            if header_idx > 0:
                # Set the real header
                new_header = df.iloc[header_idx]
                df = df[header_idx+1:]
                df.columns = new_header
                
        except Exception as e:
            print(f"Fallback to standard pandas read due to: {e}")
            # Fallback to standard pandas read
            sample_df = pd.read_excel(file_path, nrows=20, header=None)
            header_idx = find_header_row(sample_df)
            df = pd.read_excel(file_path, skiprows=header_idx)
            
        # Clean data: drop completely empty rows/cols
        df.dropna(how='all', inplace=True)
        df.dropna(axis=1, how='all', inplace=True)
        
        # Clean column names
        df.columns = clean_column_names(df.columns)
        
        # Write to SQLite
        df.to_sql(table_name, conn, index=False, if_exists='replace')
        print(f"Excel import completed successfully. Loaded {len(df)} rows.")
        
    else:
        raise ValueError(f"Unsupported file format: {ext}")
        
    conn.close()
    return table_name

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Import Excel/CSV to SQLite")
    parser.add_argument("input_file", help="Path to the input Excel or CSV file")
    parser.add_argument("--db", default="workspace.db", help="Path to SQLite database file")
    parser.add_argument("--table", default=None, help="Target table name (auto-generated if not provided)")
    
    args = parser.parse_args()
    
    try:
        final_table = import_to_sqlite(args.input_file, args.db, args.table)
        print(f"SUCCESS: Data imported into table '{final_table}'")
    except Exception as e:
        print(f"ERROR: {e}")
